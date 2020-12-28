from shutil import copyfile
from openpyxl import load_workbook


def new235():  # make a blank copy of a 235 form
    # Puts a new 235 form in complete_forms to be filled out by fill235 method
    copyfile("form_templates/235.xlsx", "complete_forms/TABC235.xlsx")


def new236():  # make a blank copy of a 236 form
    # Puts a new 236 form in complete_forms to be filled out by fill236 method
    copyfile("form_templates/236.xlsx", "complete_forms/TABC236.xlsx")


'''
Used to fill out TABC 235 (monthly report of beer)
Excel does some of the basic calculations, those are not included here.
'''


class TABCForm235:
    def __init__(self):
        # declaring all line items that need to be filled out. excel sheet does some of the math
        self.line1 = 0  # Inventory, Beginning of month
        self.line2 = 0  # Beer Manufactured
        self.line3 = 0  # Beer Imported
        self.line4 = 0  # Beer Returned from TX Distributors
        self.line6 = 0  # Inventory, End of month
        self.line7 = 0  # Distributor Sales
        self.line8 = 0  # Other Exemptions
        self.line12 = 0  # Total beer sold to Retailers
        self.line13 = 0  # Total beer sold for On-Premise consumption
        self.line14 = 0  # Total beer sold for Off-Premise consumption
        self.authorizedCredits = 0  # attach tabc letter

    def fill235(self):  # fill out a 235 form using object data
        new235()  # make a new 235 form to fill out
        book = load_workbook('form_templates/235.xlsx')
        sheet = book.active
        # fill blank lines
        sheet['B14'] = self.line1
        sheet['B15'] = self.line2
        sheet['B16'] = self.line3
        sheet['B17'] = self.line4
        sheet['B19'] = self.line6
        sheet['B20'] = self.line7
        sheet['B21'] = self.line8
        sheet['B26'] = self.line12
        sheet['B27'] = self.line13
        sheet['B28'] = self.line14

        book.save('complete_forms/TABC235.xlsx')


class TABCForm236:
    def __init__(self):
        self.line1 = 0
        self.line2 = 0
        self.line3 = 0
        self.line4 = 0
        self.line6 = 0
        self.line7 = 0
        self.line8 = 0
        self.line12 = 0
        self.line13 = 0
        self.line14 = 0
        self.authorizedCredits = 0

    def fill236(self):  # fill out a 236 form using object data
        new235()  # make a new 236 form to fill out
        book = load_workbook('form_templates/236.xlsx')
        sheet = book.active
        # fill blank lines
        sheet['B14'] = self.line1
        sheet['B15'] = self.line2
        sheet['B16'] = self.line3
        sheet['B17'] = self.line4
        sheet['B19'] = self.line6
        sheet['B20'] = self.line7
        sheet['B21'] = self.line8
        sheet['B26'] = self.line12
        sheet['B27'] = self.line13
        sheet['B28'] = self.line14

        book.save('complete_forms/TABC236.xlsx')
