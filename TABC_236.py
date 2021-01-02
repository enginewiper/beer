import pandas as pd
import numpy as np
from datetime import date, timedelta, datetime
from pandas.io import excel
from openpyxl import load_workbook
from shutil import copyfile


def new235():  # make a blank copy of a 235 form
    # Puts a new 235 form in complete_forms to be filled out by fill235 method
    currentMonth = datetime.now().month
    currentYear = datetime.now().year
    newFilename = currentYear + "_" + currentMonth + "_" + "c-235.xlsx"
    copyfile("form_templates/235.xlsx", newFilename)


def new236():  # make a blank copy of a 236 form
    # Puts a new 236 form in complete_forms to be filled out by fill236 method
    currentMonth = datetime.now().month
    currentYear = datetime.now().year
    newFilename = currentYear + "_" + currentMonth + "_" + "c-235.xlsx"
    copyfile("form_templates/236.xlsx", newFilename)


def get_previous_report_name_prefix():
    today = date.today()
    first = today.replace(day=1)
    last_month = first - timedelta(days=32)
    return last_month.strftime("%Y_%m_")


reports_rootdir = 'C:/!/projects/PycharmProjects/beer/reportsdir/'


class PreviousForm:  # Class that holds all the data from previous 235 or 236 form.
    def __init__(self, filename):  # Takes in previous form filename as parameter.
        self.workbook = load_workbook(filename, data_only=True)  # can be appended to object and use openpyxl functions
        self.inv_end_of_month = (filename['Summary Page'])['B19'].value  # Line 1
        self.beverage_brewed_ytd = (filename['Summary Page'])['C15'].value  # Line 2 YTD
        self.beverage_imported_ytd = (filename['Summary Page'])['C16'].value  # Line 3 YTD
        self.beverage_taxable_ytd = (filename['Summary Page'])['C23'].value  # Line 10 YTD
        self.tax_rate_per_gallon = (filename['Summary Page'])['B24'].value  # Line 11
        self.sold_to_retailers_ytd = (filename['Summary Page'])['C26'].value  # Line 12 YTD
        self.on_premise_consumption_ytd = (filename['Summary Page'])['C27'].value  # Line 13 YTD
        self.off_premise_consumption_ytd = (filename['Summary Page'])['C28'].value  # Line 14 YTD


previous_tabc_235_filename = reports_rootdir + get_previous_report_name_prefix() + 'c-235.xlsx'
previous_tabc_236_filename = reports_rootdir + get_previous_report_name_prefix() + 'c-236.xlsx'

prev_235_form = PreviousForm(previous_tabc_235_filename)  # Creating object using filename as parameter.

# print(previous_tabc_235_workbook.sheetnames)
# >>> ['Summary Page', 'Schedules', 'Brand Summary Sold', 'Supplemental Schedule']
print((prev_235_form.workbook['Summary Page'])['B18'].value)


# 1.   Inventory, Beginning of Month  (Line 6 on Prior Monthly Report)
inv_beginning_of_month = prev_235_form.inv_end_of_month
# 2.   Ale/Malt Liquor Brewed (Gallons Bottled & Kegged)
beverage_brewed_monthly = 0.00
# TODO calculate the amount in gallons added to inventory.xlsx during reporting period, output current_sheet b15
beverage_brewed_ytd = 0.00
# TODO sum beverage_brewed_monthly with C15 from previous ytd, output current_sheet c15


# 3.   Ale/Malt Liquor Imported (Page 2, Schedule A)
# always zero for us
beverage_imported_monthly = 0.00
beverage_imported_ytd = 0.00
# TODO output current_sheet b16 and c16

# 4.   Ale/Malt Liquor Returned from TX Wholesalers
# always zero for us
beverage_returned_from_wholesalers_monthly = 0.00
# TODO output current_sheet b17

# 5.   Total Ale/Malt Liquor Available (Sum of Lines 1,2,3,4)
beverage_total_available_monthly = inv_beginning_of_month + \
                                   beverage_brewed_monthly + \
                                   beverage_imported_monthly + \
                                   beverage_returned_from_wholesalers_monthly
# TODO output to current_sheet b18

# 6.   Inventory, End of Month
inv_end_of_month = 0.00
# TODO calculate end of month inventory from inventory.xlsx
# TODO output to current_sheet b19

# 7.   Wholesaler Sales (Page 2, line 2)
# always zero for us
wholesaler_sales_monthly = 0.00
wholesaler_sales_ytd = 0.00
# TODO output to current_sheet b20 and c20

# 8.   Other Exemptions (Page 2, line 3)
# we will calculate these manually as one-offs if there ever are any
other_exemptions_monthly = 0.00
other_exemptions_ytd = 0.00
# TODO output to current_sheet b21 and c21

# 9.   Total Exemptions (Sum of Lines 6, 7 & 8)
beverage_total_exemptions_monthly = inv_end_of_month + wholesaler_sales_monthly + other_exemptions_monthly
# TODO output to current_sheet b22

# 10.  Ale/Malt Liquor Subject to Taxation (Line 5 minus Line 9)
beverage_subject_to_taxation_monthly = beverage_total_available_monthly - beverage_total_exemptions_monthly
beverage_subject_to_taxation_ytd = (prev_235_form.beverage_taxable_ytd + beverage_subject_to_taxation_monthly)

# 11. Tax Rate Per Gallon
tax_rate_per_gal = prev_235_form.tax_rate_per_gallon

# Compliance Totals
# 12.  Total Ale/Malt Liquor Sold to Retailers (Page 3)
# TODO generate Brand Summary Sold sheet inputs from transactions.xlsx
# calculate total values monthly and ytd
# output to current_sheet B26 and C26

# 13.  Total Ale/Malt Liquor Sold for On-Premise Consumption
# TODO generate total ale/malt liquor sold on-prem from transactions.xlsx
# calculate total values monthly and ytd
# output to current_sheet B27 and C27

# 14.  Total Ale/Malt Liquor Sold for Off-Premise Consumption
# TODO generate total ale/malt liquor sold off-prem from transactions.xlsx
# calculate total values monthly and ytd
# output to current_sheet B28 and C28

# 15.  Total Taxable Sales* (Sum of Lines 12,13, & 14)

# TODO calculate total taxable sales monthly
# output to current_sheet B29

gross_tax_due = beverage_subject_to_taxation_monthly * tax_rate_per_gal
# TODO output to current_sheet B32

discount_tax_due = gross_tax_due - (gross_tax_due * .02)
# TODO output to current_sheet B33

less_authorized_credits = discount_tax_due
# TODO output to current_sheet B34

tax_due_state = discount_tax_due
# TODO output to current_sheet B35
