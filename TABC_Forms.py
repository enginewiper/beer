import pandas as pd
from datetime import date, timedelta, datetime
from pandas.io import excel
from openpyxl import load_workbook
from shutil import copyfile

# Set paths for all files statically.
invPath = 'Inventory.xlsx'
productsPath = 'Products.xlsx'
retailerCustomersPath = 'RetailerCustomers.xlsx'
transactionsPath = 'Transactions.xlsx'


def convert_to_gallons(amount, units):  # Can do G, oz, pt, qt, L
    if units == 'G' or 'g':
        return amount  # does not need to be converted
    elif units == 'oz' or 'OZ':
        return amount/128  # there are 128 oz in a gallon
    elif units == 'pt' or 'PT':
        return amount/8  # there are 8 pints in a gallon
    elif units == 'qt' or 'QT':
        return amount/4  # there are 4 quarts in a gallon
    elif units == 'L' or 'l':
        return amount/3.785  # there are 3.785 liters in a gallon
    else:
        return "ERR"  # this should never happen


def get_previous_report_name_prefix():
    today = date.today()
    first = today.replace(day=1)
    last_month = first - timedelta(days=32)
    return last_month.strftime("%Y_%m_")


reports_rootdir = 'C:/Users/svenf/Desktop/beer/reportsdir/'
# reports_rootdir = 'C:/!/projects/PycharmProjects/beer/reportsdir/'


previous_tabc_235_filename = reports_rootdir + get_previous_report_name_prefix() + 'c-235.xlsx'
previous_tabc_236_filename = reports_rootdir + get_previous_report_name_prefix() + 'c-236.xlsx'


# is a transaction date within the previous month? boolean
def is_current_reporting_period(transaction_date):
    last_day_of_prev_month = date.today().replace(day=1) - timedelta(days=1)
    start_day_of_prev_month = date.today().replace(day=1) - timedelta(days=last_day_of_prev_month.day)
    return start_day_of_prev_month <= transaction_date <= last_day_of_prev_month


def new235():  # make a blank copy of a 235 form for the current month and year
    # Puts a new 235 form in complete_forms to be filled out by fill235 method
    current_month = str(datetime.now().month)
    current_year = str(datetime.now().year)
    new_filename = "complete_forms/" + current_year + "_" + current_month + "_" + "c-235.xlsx"
    copyfile("form_templates/235.xlsx", new_filename)
    return new_filename


def new236():  # make a blank copy of a 236 form for the current month and year
    # Puts a new 236 form in complete_forms to be filled out by fill236 method
    current_month = str(datetime.now().month)
    current_year = str(datetime.now().year)
    new_filename = "complete_forms/" + current_year + "_" + current_month + "_" + "c-236.xlsx"
    copyfile("form_templates/236.xlsx", new_filename)
    return new_filename


class Transaction:
    def __init__(self,
                 product_id,
                 individual_container_size,
                 container_units,
                 number_units,
                 is_off_premise,
                 is_retailer_sale):
        self.product_id = str(product_id)  # Product ID
        self.individual_container_size = str(individual_container_size)  # Individual Container Size
        self.container_units = str(container_units)  # oz or gallons?
        self.number_units = int(number_units)  # number of containers sold
        self.total_gallons = number_units * convert_to_gallons(individual_container_size, container_units)
        self.is_off_premise = str(is_off_premise)
        self.is_retailer_sale = is_retailer_sale  # 0 or 1


class Product:
    def __init__(self, brand_name, product_id, is_liquor):  # Product name, product ID, is it liquor or not (0 or 1)?
        self.alphabetical_order = 0  # Used to find which row of the tax form that the product data will be written to.
        self.brandName = str(brand_name)
        self.productID = str(product_id)
        self.isLiquor = is_liquor
        self.ONP_total_gallons_sold = 0  # ON PREMISE SALES total gallons
        self.OFFP_total_gallons_sold = 0  # OFF PREMISE SALES
        self.RTL_sixth_barrel_sold = 0  # RETAILER SALES (add different container types below this line)
        self.RTL_twentytwo_oz_bottle_sold = 0
        self.RTL_total_gallons_sold = 0


def generate_transaction_list():  # Generate a list of Transaction objects with data needed for both forms.
    transactions = []
    df_transactions = pd.io.excel.read_excel(transactionsPath)

    for row in df_transactions.itertuples(index=False):  # iterate through all rows in dataframe
        if is_current_reporting_period(row[df_transactions.columns.get_loc('Date')]):  # Getting data into local vars
            output_product_id = row[df_transactions.columns.get_loc('ProductID')]  # Getting product ID into local var
            output_individual_container_size = row[
                df_transactions.columns.get_loc('ContainerSize')]  # Getting container size #
            output_container_units = row[
                df_transactions.columns.get_loc('ContainerUnits')]  # Getting units (oz or gallons?)
            output_number_units = row[df_transactions.columns.get_loc('NumUnits')]  # Getting number of containers sold
            is_off_premise = row[df_transactions.columns.get_loc('OffPrem')]  # is it off premise or on premise
            if is_off_premise == "TRUE":
                is_off_premise = 1
            elif is_off_premise == "FALSE":
                is_off_premise = 0
            internal_customer_id = row[df_transactions.columns.get_loc('InternalCustomerID')]  # internal customer id
            cxid = int(internal_customer_id)  # Converting internal customer ID to an int.
            is_rtl_sale = 0  # boolean, 0 or 1
            if cxid != 1:  # If internal customer ID is not 1, it's a retailer sale.
                is_rtl_sale = 1
            elif cxid == 1:
                is_rtl_sale = 0
            transactions.append(Transaction(product_id=output_product_id,
                                            individual_container_size=output_individual_container_size,
                                            container_units=output_container_units,
                                            number_units=output_number_units,
                                            is_off_premise=is_off_premise,
                                            is_retailer_sale=is_rtl_sale)
                                )
    return transactions


def test_transaction_list(transaction_list_name):
    test_list = transaction_list_name
    for x in range(0, len(test_list)):
        print("PRODUCT ID:" + test_list[x].product_id)
        print(str(test_list[x].number_units) + " of " + str(test_list[x].individual_container_size) + test_list[x].container_units)
        print("Off premise: " + test_list[x].is_off_premise)
        if test_list[x].is_retailer_sale == 1:
            print("(Retailer sale)")
        print("\n")


# puts product data into an array of objects
def generate_product_list(product_book):  # Put file path / filename for Products.xlsx as the parameter.
    book = load_workbook(product_book, data_only=True)
    product_list = []  # array that objects will go in
    number_of_products = -1  # Starts at -1 so it won't count the top cell.
    x_column = book.active['A']  # for this part it doesn't actually matter which column
    for x in x_column:
        if x.value is None:
            break
        number_of_products = number_of_products + 1

    for x in range(0, number_of_products):  # Make (number_of_products) products
        d = book.active[("D" + str(x+2))].value  # find brand name for new object
        a = book.active[("A" + str(x+2))].value  # find productID for new object
        b = book.active[("B" + str(x+2))].value  # find isLiquor for new object
        is_liquor = 0
        if b == "B":
            is_liquor = 0
        elif b == "ML":
            is_liquor = 1
        product_list.append(Product(d, a, is_liquor))

    return product_list


def test_product_list(given_array):
    test_list = given_array
    for x in range(0, len(test_list)):
        print("BRAND NAME:")
        print(test_list[x].brandName)
        print(" PRODUCT ID:")
        print(test_list[x].productID)
        print(" IS LIQUOR:")
        print(test_list[x].isLiquor)
        print("Total Gallons Sold to Retailers:" + str(test_list[x].RTL_total_gallons_sold))
        print("Sixth Barrels sold to Retailers:" + str(test_list[x].RTL_sixth_barrel_sold))
        print("22oz Bottles sold to Retailers:" + str(test_list[x].RTL_twentytwo_oz_bottle_sold))
        print("Total gallons sold On premise consumption:" + str(test_list[x].ONP_total_gallons_sold))
        print("Total gallons sold Off premise consumption:" + str(test_list[x].OFFP_total_gallons_sold))
        print("\n")


class PreviousForm:  # Class that holds all the data from previous 235 or 236 form.
    def __init__(self, filename):  # Takes in previous form filename as parameter.
        self.workbook = load_workbook(filename, data_only=True)  # can be appended to object and use openpyxl functions
        self.inv_end_of_month = (self.workbook['Summary Page'])['B19'].value  # Line 1
        self.beverage_brewed_ytd = (self.workbook['Summary Page'])['C15'].value  # Line 2 YTD
        self.beverage_imported_ytd = (self.workbook['Summary Page'])['C16'].value  # Line 3 YTD
        self.distributor_sales_ytd = (self.workbook['Summary Page'])['C20'].value # Line 7 YTD
        self.other_exemptions_ytd = (self.workbook['Summary Page'])['C21'].value  # Line 8 YTD
        self.beverage_taxable_ytd = (self.workbook['Summary Page'])['C23'].value  # Line 10 YTD
        self.tax_rate_per_gallon = (self.workbook['Summary Page'])['B24'].value  # Line 11
        self.sold_to_retailers_ytd = (self.workbook['Summary Page'])['C26'].value  # Line 12 YTD
        self.on_premise_consumption_ytd = (self.workbook['Summary Page'])['C27'].value  # Line 13 YTD
        self.off_premise_consumption_ytd = (self.workbook['Summary Page'])['C28'].value  # Line 14 YTD


class Current235:  # Class that holds data to be put in the new 235 form.
    def __init__(self):
        # new235 will return the filename after creating a new form
        self.filename = new235()
        self.wb = load_workbook(self.filename, data_only=True)
        self.p1 = self.wb["Summary Page"]
        self.p2 = self.wb["Schedules"]
        self.p3 = self.wb["Brand Summary Sold"]
        self.p4 = self.wb["Supplemental Schedule"]
        # Summary Page Variables
        self.line1 = 0  # Inventory, Beginning of month
        self.line2 = 0  # Beer Manufactured
        self.line2_YTD = 0
        self.line3 = 0  # Beer Imported
        self.line3_YTD = 0
        self.line4 = 0  # Beer Returned from TX Distributors
        self.line5 = 0  # Total Beer Available
        self.line6 = 0  # Inventory, End of month
        self.line7 = 0  # Distributor sales
        self.line7_YTD = 0
        self.line8 = 0  # Other Exemptions
        self.line8_YTD = 0
        self.line9 = 0  # Total Exemptions
        self.line10 = 0  # Beer subject to taxation
        self.line10_YTD = 0
        self.line12 = 0  # Total beer sold to retailers
        self.line12_YTD = 0
        self.line13 = 0  # Total beer sold for on-premise consumption
        self.line13_YTD = 0
        self.line14 = 0  # Total beer sold for off-premise consumption
        self.line14_YTD = 0
        self.line15 = 0  # Total taxable sales
        self.tax_rate_per_gallon = self.p1['B24'].value
        self.gross_tax = 0  # Gross Tax Due
        self.less_2percent = 0
        self.less_authorized_credits = 0
        self.tax_due_state = 0

    def get_line_5(self):
        self.line5 = self.line1 + self.line2 + self.line3 + self.line4

    def get_line_15(self):
        self.line15 = self.line12 + self.line13 + self.line14

    def calculate_gross_tax(self):
        self.gross_tax = self.line10 * self.tax_rate_per_gallon

    def calculate_total_tax(self):
        self.less_2percent = self.gross_tax * 0.02
        self.tax_due_state = self.gross_tax - self.less_authorized_credits - self.less_2percent

    def get_ytds(self, previous_form_obj):
        today = datetime.today()
        current_month = today.month
        if current_month == 2:  # If we are making the first form of the year then don't look at previous form ytd
            self.line2_YTD = self.line2
            self.line3_YTD = self.line3
            self.line7_YTD = self.line7
            self.line8_YTD = self.line8
            self.line10_YTD = self.line10
            self.line12_YTD = self.line12
            self.line13_YTD = self.line13
            self.line14_YTD = self.line14
        else:
            self.line2_YTD = previous_form_obj.beverage_brewed_ytd + self.line2
            self.line3_YTD = previous_form_obj.beverage_imported_ytd + self.line3
            self.line7_YTD = previous_form_obj.distributor_sales_ytd + self.line7
            self.line8_YTD = previous_form_obj.other_exemptions_ytd + self.line8
            self.line10_YTD = previous_form_obj.beverage_taxable_ytd + self.line10
            self.line12_YTD = previous_form_obj.sold_to_retailers_ytd + self.line12
            self.line13_YTD = previous_form_obj.on_premise_consumption_ytd + self.line13
            self.line14_YTD = previous_form_obj.off_premise_consumption_ytd + self.line14

    def fill_summary_page(self):
        self.p1['B14'] = self.line1
        self.p1['B15'] = self.line2
        self.p1['C15'] = self.line2_YTD
        self.p1['B16'] = self.line3
        self.p1['C16'] = self.line3_YTD
        self.p1['B17'] = self.line4
        self.p1['B18'] = self.line5
        self.p1['B19'] = self.line6
        self.p1['B20'] = self.line7
        self.p1['C20'] = self.line7_YTD
        self.p1['B21'] = self.line8
        self.p1['C21'] = self.line8_YTD
        self.p1['B22'] = self.line9
        self.p1['B23'] = self.line10
        self.p1['C23'] = self.line10_YTD
        self.p1['B26'] = self.line12
        self.p1['C26'] = self.line12_YTD
        self.p1['B27'] = self.line13
        self.p1['C27'] = self.line13_YTD
        self.p1['B28'] = self.line13
        self.p1['C28'] = self.line13_YTD
        self.p1['B29'] = self.line15
        self.p1['B32'] = self.gross_tax
        self.p1['B33'] = self.less_2percent
        self.p1['B34'] = self.less_authorized_credits
        self.p1['B35'] = self.tax_due_state
        self.wb.save(self.filename)


class Current236:  # Class that holds data to be put in the new 235 form.
    def __init__(self):
        # new235 will return the filename after creating a new form
        self.filename = new236()
        self.wb = load_workbook(self.filename, data_only=True)
        self.p1 = self.wb["Summary Page"]
        self.p2 = self.wb["Schedules"]
        self.p3 = self.wb["Brand Summary Sold"]
        self.p4 = self.wb["Supplemental Schedule"]
        # Summary Page Variables
        self.line1 = 0  # Inventory, Beginning of month
        self.line2 = 0  # Ale/Malt Liquor Manufactured
        self.line2_YTD = 0
        self.line3 = 0  # Ale/Malt Liquor Imported
        self.line3_YTD = 0
        self.line4 = 0  # Ale/Malt Liquor Returned from TX Distributors
        self.line5 = 0  # Total Ale/Malt Liquor Available
        self.line6 = 0  # Inventory, End of month
        self.line7 = 0  # Distributor sales
        self.line7_YTD = 0
        self.line8 = 0  # Other Exemptions
        self.line8_YTD = 0
        self.line9 = 0  # Total Exemptions
        self.line10 = 0  # Ale/Malt Liquor subject to taxation
        self.line12 = 0  # Total Ale/Malt Liquor sold to retailers
        self.line12_YTD = 0
        self.line13 = 0  # Total Ale/Malt Liquor sold for on-premise consumption
        self.line13_YTD = 0
        self.line14 = 0  # Total Ale/Malt Liquor sold for off-premise consumption
        self.line14_YTD = 0
        self.line15 = 0  # Total taxable sales
        self.tax_rate_per_gallon = self.p1['B24'].value
        self.gross_tax = 0  # Gross Tax Due
        self.less_2percent = 0
        self.less_authorized_credits = 0
        self.tax_due_state = 0

    def get_line_5(self):
        self.line5 = self.line1 + self.line2 + self.line3 + self.line4

    def get_line_15(self):
        self.line15 = self.line12 + self.line13 + self.line14

    def calculate_gross_tax(self):
        self.gross_tax = self.line10 * self.tax_rate_per_gallon

    def calculate_total_tax(self):
        self.less_2percent = self.gross_tax * 0.02
        self.tax_due_state = self.gross_tax - self.less_authorized_credits - self.less_2percent

    def get_ytds(self, previous_form_obj):
        today = datetime.today()
        current_month = today.month
        if current_month == 2:  # If we are making the first form of the year then don't look at previous form ytd
            self.line2_YTD = self.line2
            self.line3_YTD = self.line3
            self.line7_YTD = self.line7
            self.line8_YTD = self.line8
            self.line10_YTD = self.line10
            self.line12_YTD = self.line12
            self.line13_YTD = self.line13
            self.line14_YTD = self.line14
        else:
            self.line2_YTD = previous_form_obj.beverage_brewed_ytd + self.line2
            self.line3_YTD = previous_form_obj.beverage_imported_ytd + self.line3
            self.line7_YTD = previous_form_obj.distributor_sales_ytd + self.line7
            self.line8_YTD = previous_form_obj.other_exemptions_ytd + self.line8
            self.line10_YTD = previous_form_obj.beverage_taxable_ytd + self.line10
            self.line12_YTD = previous_form_obj.sold_to_retailers_ytd + self.line12
            self.line13_YTD = previous_form_obj.on_premise_consumption_ytd + self.line13
            self.line14_YTD = previous_form_obj.off_premise_consumption_ytd + self.line14

    def fill_summary_page(self):
        self.p1['B14'] = self.line1
        self.p1['B15'] = self.line2
        self.p1['C15'] = self.line2_YTD
        self.p1['B16'] = self.line3
        self.p1['C16'] = self.line3_YTD
        self.p1['B17'] = self.line4
        self.p1['B18'] = self.line5
        self.p1['B19'] = self.line6
        self.p1['B20'] = self.line7
        self.p1['C20'] = self.line7_YTD
        self.p1['B21'] = self.line8
        self.p1['C21'] = self.line8_YTD
        self.p1['B22'] = self.line9
        self.p1['B23'] = self.line10
        self.p1['C23'] = self.line10_YTD
        self.p1['B26'] = self.line12
        self.p1['C26'] = self.line12_YTD
        self.p1['B27'] = self.line13
        self.p1['C27'] = self.line13_YTD
        self.p1['B28'] = self.line13
        self.p1['C28'] = self.line13_YTD
        self.p1['B29'] = self.line15
        self.p1['B32'] = self.gross_tax
        self.p1['B33'] = self.less_2percent
        self.p1['B34'] = self.less_authorized_credits
        self.p1['B35'] = self.tax_due_state
        self.wb.save(self.filename)


array_transaction_data = generate_transaction_list()  # List of transactions in the current reporting period.
array_products = generate_product_list("Products.xlsx")  # List of products
product_id_dict = {p.productID: p for p in array_products}  # Creates a dictionary of product IDs that point to objects.
prev_235_form = PreviousForm(previous_tabc_235_filename)  # Creating object using filename as parameter.
prev_236_form = PreviousForm(previous_tabc_236_filename)
current_235_form = Current235()
current_236_form = Current236()

for x in range(0, len(array_transaction_data)):  # For the number of transactions in the array
    # Intermediary variables for total amounts sold in transaction
    total_gallons = array_transaction_data[x].total_gallons  # total gallons for this transaction
    total_22oz_bottles = 0
    total_sixth_barrels = 0
    # If the transaction is 22oz bottles
    if array_transaction_data[x].individual_container_size == "22" and array_transaction_data[x].container_units == "oz":
        total_22oz_bottles = array_transaction_data[x].number_units

    # If the transaction is sixth barrels
    if array_transaction_data[x].individual_container_size == "5.12" and array_transaction_data[x].container_units == "G":
        total_sixth_barrels = array_transaction_data[x].number_units

    # If the transaction is a retailer sale
    if array_transaction_data[x].is_retailer_sale == 1:
        # Increase product object data by amounts sold
        product_id_dict[array_transaction_data[x].product_id].RTL_sixth_barrel_sold += total_sixth_barrels
        product_id_dict[array_transaction_data[x].product_id].RTL_twentytwo_oz_bottle_sold += total_22oz_bottles
        product_id_dict[array_transaction_data[x].product_id].RTL_total_gallons_sold += total_gallons

    # If the transaction is not a retailer sale
    if array_transaction_data[x].is_retailer_sale == 0:
        if array_transaction_data[x].is_off_premise == "True":
            product_id_dict[array_transaction_data[x].product_id].OFFP_total_gallons_sold += total_gallons

        if array_transaction_data[x].is_off_premise == "False":
            product_id_dict[array_transaction_data[x].product_id].ONP_total_gallons_sold += total_gallons

# Create a dataframe from inventory spreadsheet
df_inventory = pd.io.excel.read_excel(invPath)

# Write all static data to current 235 form object
current_235_form.line1 = prev_235_form.inv_end_of_month

# Write all static data to current 236 form object
current_236_form.line1 = prev_236_form.inv_end_of_month

# Find how much of each beverage was manufactured / left inventory, sort between 235/236 forms
beer_left_inventory = 0  # How much beer was sold based on inventory.xlsx?
liquor_left_inventory = 0  # How much liquor was sold based on inventory.xlsx?

for row in df_inventory.itertuples(index=False):  # iterate through all rows in inventory dataframe,
    date = row[df_inventory.columns.get_loc('PackageDate')]  # all these vars are being set to make the logic readable
    status = int(row[df_inventory.columns.get_loc('Status')])
    container_size = row[df_inventory.columns.get_loc('ContainerSize')]
    container_units = row[df_inventory.columns.get_loc('ContainerUnit')]
    contents = row[df_inventory.columns.get_loc('Contents')]
    internal_customer_id = row[df_inventory.columns.get_loc('InternalCustomerID')]

    if is_current_reporting_period(date):  # ,in the reporting period
        if status == 1:  # If the keg status is 1, it was manufactured.
            if product_id_dict[contents].isLiquor == 1:  # If it is malt liquor
                current_236_form.line2 += convert_to_gallons(container_size, container_units)  # add # gallons to 236
            if product_id_dict[contents].isLiquor == 0:  # If it is beer
                current_235_form.line2 += convert_to_gallons(container_size, container_units)  # add # gallons to 235
        if status == 2:  # If the keg status is 2, it was sold and is no longer in inventory.
            if product_id_dict[contents].isLiquor == 1:  # If it is malt liquor
                liquor_left_inventory += convert_to_gallons(container_size, container_units)  # put in local var
            if product_id_dict[contents].isLiquor == 0:  # If it is beer
                beer_left_inventory += convert_to_gallons(container_size, container_units)  # put in local var

# Post-calculation of other lines
current_235_form.get_line_5()  # add line 1+2+3+4
current_236_form.get_line_5()  # add line 1+2+3+4
current_235_form.line6 -= beer_left_inventory  # Line 6 is total beer available minus beer that was sold
current_236_form.line6 -= liquor_left_inventory  # Line 6 is total liquor available minus liquor that was sold
current_235_form.line10 = current_235_form.line5 - current_235_form.line9  # Beer subject to taxation
current_236_form.line10 = current_236_form.line5 - current_236_form.line9  # Liquor subject to taxation

# Get total taxable sales from each product, put them in the correct form.
for x in range(0, len(array_products)):
    if array_products[x].isLiquor == 0:
        current_235_form.line12 += array_products[x].RTL_total_gallons_sold
        current_235_form.line13 += array_products[x].ONP_total_gallons_sold
        current_235_form.line14 += array_products[x].OFFP_total_gallons_sold
    if array_products[x].isLiquor == 1:
        current_236_form.line12 += array_products[x].RTL_total_gallons_sold
        current_236_form.line13 += array_products[x].ONP_total_gallons_sold
        current_236_form.line14 += array_products[x].OFFP_total_gallons_sold

current_235_form.get_line_15()  # Get total taxable sales for 235 form
current_236_form.get_line_15()  # Get total taxable sales for 235 form
current_235_form.calculate_gross_tax()  # Calculate gross tax for 235 form
current_235_form.calculate_total_tax()  # Calculate total tax for 235 form
current_236_form.calculate_gross_tax()  # Calculate gross tax for 236 form
current_236_form.calculate_total_tax()  # Calculate total tax for 236 form

# Calculate YTDs
current_235_form.get_ytds(prev_235_form)
current_236_form.get_ytds(prev_236_form)

# Output data to 235 form page 1
current_235_form.fill_summary_page()
# Output data to 235 form page 3

# Output data to 236 form page 1
current_235_form.fill_summary_page()
# Output data to 236 form page 3
